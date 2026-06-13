import io
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _build_pdf_response(title, headers, rows):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(
        f"ARMADA DE REPÚBLICA DOMINICANA - Taller de Electrónica",
        styles['Heading1']
    ))
    elements.append(Paragraph(title, styles['Heading2']))
    elements.append(Paragraph(
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 20))

    data = [headers] + rows
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def _build_excel_response(title, headers, rows):
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.merge_cells('A1:' + chr(64 + len(headers)) + '1')
    ws['A1'] = f"ARMADA DE REPÚBLICA DOMINICANA - Taller de Electrónica"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:' + chr(64 + len(headers)) + '2')
    ws['A2'] = title
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:' + chr(64 + len(headers)) + '3')
    ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].alignment = Alignment(horizontal='center')

    header_row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, header_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_report(title, headers, rows, format='pdf'):
    if format == 'pdf':
        return _build_pdf_response(title, headers, rows)
    elif format == 'excel':
        return _build_excel_response(title, headers, rows)
    raise ValueError("Format must be 'pdf' or 'excel'")
